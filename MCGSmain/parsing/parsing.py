#!/usr/bin/python3
import binascii
import openpyxl
from openpyxl.cell import get_column_letter
import re
import math
from bitstring import BitArray



class ParseInfo:
    def __init__(self, sheet, nameInd, byteInd, typeInd, convInd, unitInd):
        self.sheet = sheet
        self.varNames = self.getColVals(nameInd)
        tempByteVals = self.getColVals(byteInd)
        self.typeVals = self.getColVals(typeInd)
        self.convVals = self.getColVals(convInd)
        self.unitVals = self.getColVals(unitInd)

        self.byteVals = []

        # Want byte values as integers
        for item in tempByteVals:
            self.byteVals.append(int(item)*2) # One byte is two hex values

    def getColVals(self, indices):
        listOut = []
        for row in self.sheet[indices[0]:indices[1]]:
            for obj in row:
                listOut.append(obj.value)
        return listOut
    

class Reader:
    maxRow = 0
    maxCol = 0

    # Regex for exponential notation. 
    expRe = re.compile('\d+(\.\d+)? +(e|E)? +(\d+)?')

    def __init__(self, xlsxFileName):
        wb = openpyxl.load_workbook(xlsxFileName)

        sheetList = wb.get_sheet_names()

        self.sheet = wb.get_sheet_by_name(sheetList[0])

        self.maxRow = self.sheet.max_row

        self.maxCol = self.sheet.max_column

        firstRowCoords = tuple(self.sheet['A1':get_column_letter(self.maxCol)+'1'])

        self.firstRow = []
        
        for obj in firstRowCoords:
            for temp in obj:
                self.firstRow.append(temp.value)
                
        self.nameColInd = self.findIndex('field')
        self.byteColInd = self.findIndex('byt')
        self.typeColInd = self.findIndex('type')
        self.convColInd = self.findIndex('conv')
        self.unitColInd = self.findIndex('units')

        self.parsingInfo = ParseInfo(self.sheet, self.nameColInd, self.byteColInd, self.typeColInd, self.convColInd, self.unitColInd)


    def findIndex(self, searchStr):
        # Finds the indices of a column in the telemetry definition
        colVals = [item for item in self.firstRow if searchStr in str(item).lower()]
        minIndex = chr(ord(get_column_letter(self.firstRow.index(colVals[0]) + 1))) + str(2)
        maxIndex = chr(ord(get_column_letter(self.firstRow.index(colVals[0]) + 1))) + str(self.maxRow)
        # Cell indices start at one in openpyxl, lists start at 0 in Python.  +1 is to adjust
        if not colVals:
            print("Something in the telemetry definition is off. See telemetryGuide.txt for conventions")
        return minIndex, maxIndex

    def splitter(self, inFile):
        # Reads in hex data from file, converts to binary, returns list 
        f = open(inFile)
        raw = f.read()
        f.close()

        total = 0 # Keeping track of where we are in the file

        binVals = []
        
        for length in self.parsingInfo.byteVals:
            chunk = raw[total:(total+length)]
            total += length
            numBits = int(len(chunk) * 4.0) # Defining padding 
            binVal = bin(int(chunk, 16))[2:].zfill(numBits)
            binVals.append(binVal)

        return binVals

    def parsing(self, binVals):
        # Uses binary data and converts back into readable
        # Not all decimal values, some strings
        decVals = []
        
        for binVal, type, conv in zip(binVals,self.parsingInfo.typeVals,self.parsingInfo.convVals):
            binArray = BitArray(bin=binVal)
            if 'int' in type: # Narrows it down to uint/int
                if 'u' in type:
                    decVals.append(binArray.uint)
                else:
                    decVals.append(binArray.int)
            elif 'doub' in type or 'float' in type:
                decVals.append(binArray.float)
            elif 'bool' in type:
                decVals.append(binArray.uint) # Should just be 0 or 1
            else:
                decVals.append(binArray.bytes.decode('utf-8'))
        return decVals

    def conversion(self, decVals):
        # Applies engineering unit conversions

        convVals = []
        
        for decVal, conv in zip(decVals,self.parsingInfo.convVals):
            if conv == 'NaN':
                convVals.append(decVal) # No conversion
            else:
                mo = self.expRe.search(conv)
                if not mo:
                    raise Exception('Weird value in unit conversions')
                else:
                    convVals.append(float(mo.group()) * decVal)
        return list(zip(convVals,self.parsingInfo.unitVals))

def main(telemetryDef, parseQueue, databaseQueue):
    generalReader = Reader(telemetryDef) # Create instance of reader class
    # This instance can be used for all future calls

    while True:
        if not parseQueue.empty():
            inFile = parseQueue.get()
            binVals = generalReader.splitter(inFile)
            decVals = generalReader.parsing(binVals)
            convVals = generalReader.conversion(decVals)

            databaseQueue.put(convVals)
    

    
                
        
