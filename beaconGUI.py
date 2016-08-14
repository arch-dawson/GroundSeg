import tkinter as tk
import openpyxl
from openpyxl.cell import get_column_letter
import itertools
import re

# TODO Have this initialization stuff in a class, only created once
# TODO Add a way to update information with most recent beacon 
    # Pull from database? Could show that whole chain works

def getColVals(indices):
    listOut = []
    for row in sheet[indices[0]:indices[1]]:
        for obj in row:
            listOut.append(obj.value)
    return listOut

telemetryDef = 'MCGSmain/BeaconDefinition.xlsx'

wb = openpyxl.load_workbook(telemetryDef)
sheetList = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetList[0])
maxRow = sheet.max_row
maxCol = sheet.max_column

firstCol = getColVals(('A2','A'+str(maxRow)))

firstCol.sort()

# Grouping all variables by the first 4 letters
headers = {k: list(v) for k, v in itertools.groupby(firstCol, key=lambda x: x[0:4])}

groups = []

root = tk.Tk()

headerRe = re.compile('[\W_]+') # Regex removes underscore from first 4 chars, ex: GPS_ -> GPS

handles = {}

for header in headers.keys():
    groups.append(tk.LabelFrame(root, text=headerRe.sub('',header), padx=5, pady=0))
    groups[-1].grid(row=0, column=len(groups)-1,padx=5)
    items = []
    for item in headers[header]:
        items.append(tk.Label(groups[-1],text=item))
        items[-1].pack()
    handles[header] = items

tk.mainloop()
