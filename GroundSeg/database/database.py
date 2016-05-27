#!/usr/bin/python3

import pymysql.cursors

import openpyxl

from openpyxl.cell import get_column_letter

class Database:
    def __init__(self,dbName,tableName,telemetryDef):
        self.tableName = tableName
        # Creating mysql connection
        self.dbConnection = pymysql.connect(host="localhost",
                            user="root",
                            password="P0l@r3ubE",
                            db=dbName)

        # Creating cursor to execute lines 
        self.dbCursor = self.dbConnection.cursor()

        # Checking if table exists already
        self.dbCursor.execute("SHOW TABLES LIKE '" + tableName +"';")

        # Want to run this no matter what since it makes pyxl stuff
        tableCreateStr = self.createTable(telemetryDef)

        # T/F If the table exists
        dbResult = self.dbCursor.fetchone()
        
        if not dbResult:
            self.dbCursor.execute(tableCreateStr)
            self.dbConnection.commit()
        

    def dbWrite(self, data):
        dbString = self.writeString(data)

        self.dbCursor.execute(dbString)
        self.dbConnection.commit()
        return
        

    def writeString(self, data):
        outStr = "INSERT INTO " + self.tableName + " ("
        for temp in self.firstCol:
            outStr = outStr + temp + ', '
            
        outStr = outStr[:-2]
        outStr += ") VALUES ("
        for point in data:
            if not isinstance(point, str):
                outStr = outStr + str(point) + ', '
            else:
                outStr = outStr + '\'' + str(point) + '\'' + ', '
        outStr = outStr[:-2]
        outStr += ");"

        return outStr

    def createTable(self, telemetryDef):
        wb = openpyxl.load_workbook(telemetryDef)
        sheetList = wb.get_sheet_names()
        self.sheet = wb.get_sheet_by_name(sheetList[0])
        self.maxRow = self.sheet.max_row
        self.maxCol = self.sheet.max_column

        # ==== SETUP ====
        firstRowCoords = tuple(self.sheet['A1':get_column_letter(self.maxCol)+'1'])
        
        self.firstRow = []
        
        for rowObj in firstRowCoords:
            for rowVal in rowObj:
                self.firstRow.append(rowVal.value)

        self.firstCol = self.getColVals(('A2','A'+str(self.maxRow)))

        self.typeCol = self.typeSetup()
        
        # ==== DONE SETUP ==== 

        dbCallStr = self.dbString()

        return dbCallStr
        
    def dbString(self):
        strOut = 'CREATE TABLE ' + self.tableName+ ' ('
        for first, type in zip(self.firstCol, self.typeCol):
            strOut = strOut + first + type + ', '

        strOut = strOut[:-2]  # One extra space and comma

        strOut = strOut + ');'

        return strOut

    def typeSetup(self):
        typeIndices = self.findIndex('type')
        typeVals = self.getColVals(typeIndices)

        typeList = []

        for type in typeVals:
            if 'int' in type:
                typeList.append(" BIGINT")
            elif 'bool' in type:
                typeList.append(" BOOL")
            elif 'doub' in type or 'float' in type:
                typeList.append(" DOUBLE")
            else:
                typeList.append(" TEXT(30)")

        return typeList

    def getColVals(self,indices):
        listOut = []
        for row in self.sheet[indices[0]:indices[1]]:
            for obj in row:
                listOut.append(obj.value)
        return listOut

    def findIndex(self, searchStr):
        # Finds the indices of a column in the telemetry definition
        colVals = [item for item in self.firstRow if searchStr in str(item).lower()]
        minIndex = chr(ord(get_column_letter(self.firstRow.index(colVals[0]) + 1))) + str(2)
        maxIndex = chr(ord(get_column_letter(self.firstRow.index(colVals[0]) + 1))) + str(self.maxRow)
        # Cell indices start at one in openpyxl, lists start at 0 in Python.  +1 is to adjust
        if not colVals:
            print("Something in the telemetry definition is off. See telemetryGuide.txt for conventions")
        return minIndex, maxIndex

def main(dbQueue, dbName, tableName, telemetryDef):
# During final set-up with GUI, have message box for password

    # Initializing will create table if it doesnt exist
    sqlDB = Database(dbName, tableName, telemetryDef)

    while True:
        if not dbQueue.empty():
            data, units = zip(*dbQueue.get()) #Unzips contents
            sqlDB.dbWrite(data)
        
    
