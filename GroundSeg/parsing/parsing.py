#!/usr/bin/python3
import binascii
import openpyxl
from openpyxl.cell import get_column_letter
import re
import math
import sys
from bitstring import BitArray

class Column:
    # Includes variables 
    def __init__(self, searchStr, firstRow):
        colVals = [item for item in firstRow if searchStr in str(item).lower()]
        self.minIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]))) + 1) + str(2)
        self.maxIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]))) + 1) + str(maxRow)
        if not colVals:
            print("Something in the telemetry definition is off. See telemetryGuide.txt for conventions")

class Reader:
    maxRow = 0
    maxCol = 0
    
    strRe = re.compile('str')
    intRe = re.compile('int')
    doubleRe = re.compile('doub')
    boolRe = re.compile('bool')

    def __init__(self, xlsxFileName):
        wb = openpyxl.load_workbook(xlsxFileName)

        sheetList = wb.get_sheet_names()

        self.sheet = wb.get_sheet_by_name(sheetList[0])

        self.maxRow = self.sheet.max_row

        self.maxCol = self.sheet.max_column

        firstRowCoords = tuple(self.sheet['A1':get_column_letter(self.maxCol)+'1'])

        self.firstRow = []
        
        for obj in firstRowCoords:
            for temp in obj:
                self.firstRow.append(temp.value)
                
        self.nameCol = Column('field', self.firstRow)
        #self.byteCol = Column('byt', self.firstRow)
        #self.typeCol = Column('type', self.firstRow)
        #self.convCol = Column('conv', self.firstRow)


def main(telemetryDef, parseQueue):
    generalReader = Reader(telemetryDef)
    
    
                
        
