#!/usr/bin/python3

import binascii  # Used for converting hex string to hexadecimal
import openpyxl  # Used for reading/writing excel files
from openpyxl.cell import get_column_letter 
import re  # Used for parsing unit conversions
import math  # Need to use log2 for conversions
import sys # For command line arguments
from bitstring import BitArray
import MySQLdb as sql  # For writing information to database

# Going to need to initialize all this stuff only once
# Then run main with existing indices and things



def initialize(xlsxFileName):
    
    global sheet

    global firstRow
    
    wb = openpyxl.load_workbook(xlsxFileName)  

    sheetList = wb.get_sheet_names()  # Loads in the names of the sheets

    sheet = wb.get_sheet_by_name(sheetList[0]) 

    maxRow = sheet.max_row

    maxCol = sheet.max_column

    firstRowCoords = tuple(sheet['A1':get_column_letter(maxCol)+'1'])

    firstRow = []

    for obj in firstRowCoords:
        for temp in obj:
            firstRow.append(temp.value)

    return

def indexFind(searchStr):
    try:
        colVals = [item for item in firstRow if searchStr in str(item).lower()]
        minIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]))) + 1) + str(2)
        maxIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]))) + 1) + str(maxRow)
    except:
        print("Something in the telemetry definition is off. See telemetryGuide.txt for conventions")
    return minIndex, maxIndex

def generalReader(genData):
    lengths = []
    (minIndex, maxIndex) = indexFind('byt')
    for rowOfCellObjects in sheet[minIndex:maxIndex]:  # Going through the row of values.
        for cellObj in rowOfCellObjects:  # Iterating through each cell of the row
            readVal = int(cellObj.value)
            readVal *= 2 # Bytes are two hex characters
            lengths.append(readVal)  # Appends all the lengths read in from the excel file

    total = 0  # Use this to keep track of where we are in the total number of bytes.

    binVals = []

    for length in lengths:  # Converts all of the data into binary
        chunk = data[total:(total+length)]
        total += length
        numBits = int(len(chunk) * 4.0)  # Defining the padding at the end of binary.
        binVal = bin(int(chunk, 16))[2:].zfill(numBits)
        binVals.append(binVal)

    dataCount = 0;

    decVals = [] 

    (typeMinIndex, typeMaxIndex) = indexFind('type')
    for dataConvRow in sheet[typeMinIndex:typeMaxIndex]:
        for dataObj in dataConvRow:
            binArray = BitArray(bin=binVals[dataCount])
            if intRe.search(dataObj.value): # Narrows it down to uint/int
                if 'u' in dataObj.value: # Checking if it's unsigned.  Elegant? No. Effective? Yes.
                    decVals.append(binArray.uint)
                else:
                    decVals.append(binArray.int)
            elif doubleRe.search(dataObj.value):
                decVals.append(binArray.float)
            elif boolRe.search(dataObj.value):
                decVals.append(binArray.uint) # This value should just be 0 or 1
            #else:
                #print(binVals[dataCount])
                #decVals.append(binArray.bytes)
                # 'allstarcosgc' is 12 characters long.  Imma kill someone
            dataCount += 1

    # Now we have the full set of binary data, we can re-chop it up.  This is necessary to account for partial bytes.

    convVals = []

    p = re.compile('\d+(\.\d+)? +(e|E)? +(\d+)?')  # Regular expression for number, (optional decimal and exponent)

    (unitMinIndex, unitMaxIndex) = indexFind('conv')
    for convRow in sheet[unitMinIndex:unitMaxIndex]:
        for convObj in convRow:
            if convObj.value == 'NaN':
                convVals.append(1)  # We're going to be multiplying by this, so just a 1 is fine for no conversion
            elif not p.search(convObj.value):
                raise Exception('Weird value in unit conversions.  Use \'NaN\' if no conversion, ')
            else:
                mo = p.search(convObj.value)
                convVals.append(float(mo.group()))

    # Use [a*b for a,b in zip(lista,listb)] for converting
    convertedVals = [a*b for a,b in zip(decVals,convVals)]
    return convertedVals


def dbWriteStr(dataList, tableName):
    nameStr = '('
    (nameMinIndex, nameMaxIndex) = findIndex('field')
    for nameRow in sheet[nameMinIndex:nameMaxIndex]:
        for nameValue in nameRow:
            nameStr = nameStr + nameValue + ', '
    nameStr = nameStr[:-2] + ')' # One extra space and comma
    
    valuesStr = "VALUES ("
    
    for dataVal in dataList:
        valuesStr = valuesStr + dataVal + ', '

    valuesStr = valuesStr[:-2] + ')'

    dbString = "INSERT INTO " + tableName + nameStr + valuesStr

    return dbString


def main(telemetryDef, ):
    global sheet

    # SUB IN YOUR FILE NAME HERE WHERE THE BYTE LENGTHS ARE
xlsxFileName = 'BeaconDefinition.xlsx'

# CHANGE THE NAME OF THE FILE YOU WANT TO PARSE
parseFileName = 'mostRecent'

f = open(parseFileName)  # Opens the hex beacon file to be read in.

data = f.read()
f.close()
