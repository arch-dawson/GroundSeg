#!/usr/bin/python3
import binascii
import openpyxl
from openpyxl.cell import get_column_letter
import re
import math
from bitstring import BitArray

class Column:
    # Includes variables 
    def __init__(self, searchStr, firstRow, maxRow):
        colVals = [item for item in firstRow if searchStr in str(item).lower()]
        self.minIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]) + 1))) + str(2)
        self.maxIndex = chr(ord(get_column_letter(firstRow.index(colVals[0]) + 1))) + str(maxRow)
        # Cell indices start at one in openpyxl, lists start at 0 in Python.  +1 is to adjust
        if not colVals:
            print("Something in the telemetry definition is off. See telemetryGuide.txt for conventions")

class Reader:
    maxRow = 0
    maxCol = 0
    
    strRe = re.compile('str')
    intRe = re.compile('int')
    doubleRe = re.compile('doub')
    boolRe = re.compile('bool')

    def __init__(self, xlsxFileName):
        wb = openpyxl.load_workbook(xlsxFileName)

        sheetList = wb.get_sheet_names()

        self.sheet = wb.get_sheet_by_name(sheetList[0])

        self.maxRow = self.sheet.max_row

        self.maxCol = self.sheet.max_column

        firstRowCoords = tuple(self.sheet['A1':get_column_letter(self.maxCol)+'1'])

        self.firstRow = []
        
        for obj in firstRowCoords:
            for temp in obj:
                self.firstRow.append(temp.value)
                
        self.nameCol = Column('field', self.firstRow, self.maxRow)
        self.byteCol = Column('byt', self.firstRow, self.maxRow)
        self.typeCol = Column('type', self.firstRow, self.maxRow)
        self.convCol = Column('conv', self.firstRow, self.maxRow)
        self.unitCol = Column('units', self.firstRow, self.maxRow)

    def readIn(self, inFile):
        # Input: File
        # Function: Reads in data, converts to binary
        # Output: Binary data
        f = open(inFile)
        data = f.read()

        # print(typeCol)

        
def main(telemetryDef, parseQueue):
    generalReader = Reader()
    generalReader.readIn('testFile')
