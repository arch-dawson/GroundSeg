#!/usr/bin/python3

import binascii  # Used for converting hex string to hexadecimal

import openpyxl  # Used for reading/writing excel files
from openpyxl.cell import get_column_letter 

import re  # Used for parsing unit conversions

import math  # Need to use log2 for conversions

import sys # For command line arguments

from bitstring import BitArray

# import datetime  # Used for creating the date stamp on the files


<<<<<<< HEAD
# ====== CHANGE TO REFLECT LAYOUT OF TELEMETRY DEFINITION ======
=======
# ====== CHANGE THIS BIT IF YOU CHANGE TELEMETRY DEFINITION ======
>>>>>>> master

# SUB IN YOUR FILE NAME HERE WHERE THE BYTE LENGTHS ARE
xlsxFileName = 'BeaconDefinition.xlsx'

# CHANGE THE NAME OF THE FILE YOU WANT TO PARSE
parseFileName = 'mostRecent'

# NUMBER OF ROWS AT THE TOP WITHOUT DATA
descRows = 1

# ====== END OF USER CHANGES ======

wb = openpyxl.load_workbook(xlsxFileName)  # Loads the beacon definition excel file

sheetList = wb.get_sheet_names()  # Loads in the names of the sheets.

sheet = wb.get_sheet_by_name(sheetList[0])  # As of 3/8/16 we only want the first sheet

maxRow = sheet.max_row

maxCol = sheet.max_column

firstRowCoords = tuple(sheet['A1':get_column_letter(maxCol)+'1'])

firstRow = []

for obj in firstRowCoords:
    for temp in obj:
        firstRow.append(temp.value)

# MIN AND MAX OF INDICES WHERE THE BYTE LENGTHS ARE

byteColVal = [item for item in firstRow if 'byt' in str(item).lower()]
if byteColVal:
    minIndex = chr(ord(get_column_letter(firstRow.index(byteColVal[0]))) + 1) + str(1 + descRows)
    maxIndex = chr(ord(get_column_letter(firstRow.index(byteColVal[0]))) + 1) + str(maxRow)
else:
    minIndex = 'F' + str(1 + descRows)
    maxIndex = 'F' + str(maxRow)
    
# ADD INDICES FOR UNIT CONVERSIONS LINE
# USE 'E' OR 'e' FOR EXPONENTIAL NOTATION: '123e7', NOT '123*10^7'
# USE 'NaN' OR '1' IF NO CONVERSION
unitColVal = [item for item in firstRow if 'conv' in str(item).lower()]
if byteColVal:
    unitMinIndex = chr(ord(get_column_letter(firstRow.index(unitColVal[0]))) + 1) + str(1 + descRows)
    unitMaxIndex = chr(ord(get_column_letter(firstRow.index(unitColVal[0]))) + 1) + str(maxRow)
else:
    unitMinIndex = 'H' + str(1 + descRows)
    unitMaxIndex = 'H' + str(maxRow)

# INDICES FOR DATA TYPE 
typeColVal = [item for item in firstRow if 'type' in str(item).lower()]
if byteColVal:
    typeMinIndex = chr(ord(get_column_letter(firstRow.index(typeColVal[0]))) + 1) + str(1 + descRows)
    typeMaxIndex = chr(ord(get_column_letter(firstRow.index(typeColVal[0]))) + 1) + str(maxRow)
else:
    typeMinIndex = 'G' + str(1 + descRows)
    typeMaxIndex = 'G' + str(maxRow)


# --------------------------------------------------------

f = open(parseFileName)  # Opens the hex beacon file to be read in.

data = f.read()
f.close()


### Pass only the raw data w/o header/footer?  Could remove some if statements
##def beaconReader(beaconData, beaconSheet):
##    beacLengths = []
##    for beaconRow in beaconSheet[minIndex:maxIndex]:  # Going through the row of values.
##        for cellBeacObj in beaconRow:  # Iterating through each cell of the row
##            readBeacVal = int(cellBeacObj.value)
##            beacLengths.append(readBeacVal)  # Appends all the lengths read in from the excel file
##
##    beacTotal = 0  # Use this to keep track of where we are in the total number of bytes.
##
##    outVals = []
##
##    for beacLength in beacLengths:
##        beacChunk = data[beacTotal:(beacTotal+beacLength-1)]
##        beacTotal += beacLength
##        if beacChunk == 'allstarcosgc':
##            outVals.append(beacChunk)
##            continue
##        else:
##            hexBeacVal = binascii.hexlify(beacChunk)
##            decBeacVal = int(hexBeacVal, 16)
##            outVals.append(decBeacVal)
##    return

"""
It's better to use the general parser even for beacons, since it's easier to handle uint with binary
Will need to add functionality to the program calling this to check if the file is a beacon or not.
Could also add that here.
"""

def generalReader(genData, genSheet):
    lengths = []
    for rowOfCellObjects in sheet[minIndex:maxIndex]:  # Going through the row of values.
        for cellObj in rowOfCellObjects:  # Iterating through each cell of the row
            readVal = int(cellObj.value)
            readVal *= 2 # Bytes are two hex characters
            lengths.append(readVal)  # Appends all the lengths read in from the excel file

    total = 0  # Use this to keep track of where we are in the total number of bytes.

    binVals = []

    for length in lengths:  # Converts all of the data into binary
        chunk = data[total:(total+length)]
        total += length
        numBits = int(len(chunk) * 4.0)  # Defining the padding at the end of binary.
        binVal = bin(int(chunk, 16))[2:].zfill(numBits)
        binVals.append(binVal)

#    for length in lengths:
#        length *= 8.0  # Bytes to bits
#        if not round(length) == length:  # Checking if all the bits are whole numbers
#            raise Exception('Non-whole number of bits')
#        else:
#            length = int(length)

    # JUST MAKE SURE THAT THE DATA TYPES HAVE THESE IN THEM
    # MAKE SURE THERE'S A 'u' IN THE STRING IF IT'S UINT
    strRe = re.compile('str')
    intRe = re.compile('int')
    doubleRe = re.compile('doub')
    boolRe = re.compile('bool')

    dataCount = 0;

    decVals = [] # TODO: Format output into a list of tuples, including value and type 
    
    for dataConvRow in sheet[typeMinIndex:typeMaxIndex]:
        for dataObj in dataConvRow:
            binArray = BitArray(bin=binVals[dataCount])
            if intRe.search(dataObj.value): # Narrows it down to uint/int
                if 'u' in dataObj.value: # Checking if it's unsigned.  Elegant? No. Effective? Yes.
                    decVals.append(binArray.uint)
                else:
                    decVals.append(binArray.int)
            elif doubleRe.search(dataObj.value):
                decVals.append(binArray.float)
            elif boolRe.search(dataObj.value):
                decVals.append(binArray.uint) # This value should just be 0 or 1
            #else:
                #print(binVals[dataCount])
                #decVals.append(binArray.bytes)
                # 'allstarcosgc' is 12 characters long.  Imma kill someone
            dataCount += 1

    # Now we have the full set of binary data, we can re-chop it up.  This is necessary to account for partial bytes.

    convVals = []

    p = re.compile('\d+(\.\d+)? +(e|E)? +(\d+)?')  # Regular expression for number, optional decimal and exponent

    for convRow in sheet[unitMinIndex:unitMaxIndex]:
        for convObj in convRow:
            if convObj.value == 'NaN':
                convVals.append(1)  # We're going to be multiplying by this, so just a 1 is fine for no conversion
            elif not p.search(convObj.value):
                raise Exception('Weird value in unit conversions.  Use \'NaN\' if no conversion, ')
            else:
                mo = p.search(convObj.value)
                convVals.append(float(mo.group()))

    # Use [a*b for a,b in zip(lista,listb)] for converting
    convertedVals = [a*b for a,b in zip(decVals,convVals)]
    print(convertedVals)


generalReader(data, sheet)


"""
now = datetime.datetime.now()

dateString = str(now.month) + "." + str(now.day) + "." + str(now.year)

outFileName = dateString + "." + str(outVals[timeLocation])

print outFileName
"""
