import openpyxl
import sys
from openpyxl import Workbook, load_workbook

# Used for combining all of the sheets in an excel spreadsheet into one
# Example:
#   Have test.xlsx, want output file to be named outTest.xlsx
#   python3 combineXlsx.py test.xlsx outTest.xlsx

try:
    inFile = sys.argv[1]
    outFile = sys.argv[2]
except InvalidInput:
    print("Format: combineXlsx.py inFile outFile")

wb = load_workbook(inFile)

sheets = wb.worksheets

newWS = wb.create_sheet(0)

newWS.title = "Combined"

wsIndex = 1

for sheet in sheets:
    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            newWS.cell(row=wsIndex, column=j).value = sheet.cell(row=i, column=j).value
        wsIndex += 1
    
wb.save(outFile)
